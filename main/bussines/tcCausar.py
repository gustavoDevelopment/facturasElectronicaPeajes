import json
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
from plantilla.cabecera import Cabecera
from plantilla.constants import Constants

def crear_archivo_excel_con_cabecera(base_dir,subFolder,tenant_id):
    nombre_archivo = f"documento_{subFolder}.xlsx"
    # Establecer la ruta del archivo Excel
    path_excel = os.path.join(base_dir,Constants.APLICATION_NAME.value[0],tenant_id, "output", nombre_archivo)  # Aquí creas el directorio "file"
    # Verifica que la carpeta "file" exista, si no, créala
    if not os.path.exists(os.path.dirname(path_excel)):
        os.makedirs(os.path.dirname(path_excel))

    wb = Workbook()
    ws = wb.active
    ws.title = str(Constants.FACTURA.value[0])

    for columna in Cabecera:
        ws.cell(row=1, column=columna.value[1], value=columna.value[0])

    ws2 = wb.create_sheet(title=str(Constants.NOTA_CREDITO.value[0]))
    for columna in Cabecera:
        ws2.cell(row=1, column=columna.value[1], value=columna.value[0])

    wb.save(path_excel)
    print(f"✅ Archivo creado: {path_excel}")
    return path_excel


def agregar_filas_al_excel(path_excel: str, datos: list[dict]):
    if not os.path.exists(path_excel):
        raise FileNotFoundError("El archivo no existe. Primero crea el archivo con cabeceras.")

    wb = load_workbook(path_excel)
    
    next_row = 2

    for item in datos:
        if item["InvoiceType"]== Constants.FACTURA.value[0]:
            ws = wb[str(Constants.FACTURA.value[0])]
            agregar_fila_excel(ws,item)            
        else:
            ws = wb[str(Constants.NOTA_CREDITO.value[0])]
            agregar_fila_excel(ws,item)
        
        wb.save(path_excel)
    print(f"Se agregaron {len(datos)} filas al archivo.")


# Método para agregar una fila de datos al archivo Excel
def agregar_fila_excel(ws, item):
    # Detectar la siguiente fila disponible
    next_row = ws.max_row + 1

    # Agregar datos utilizando las constantes y el item
    ws.cell(row=next_row, column=Constants.ENCAB_EMPRESA.value[1], value=Constants.ENCAB_EMPRESA.value[0])

    # Condición para determinar el tipo de documento
    if item["FacturaCabecera"] == "PP" or item["FacturaCabecera"] == "PR":
        ws.cell(row=next_row, column=Constants.ENCAB_TIPO_DOCUMENTO_FC.value[1], value=Constants.ENCAB_TIPO_DOCUMENTO_FC.value[0])
        ws.cell(row=next_row, column=Cabecera.ENCAB_NO_DTO_EXT.value[1], value=item["FacturaNumero"])
    else:
        ws.cell(row=next_row, column=Constants.ENCAB_TIPO_DOCUMENTO_NCDOC.value[1], value=Constants.ENCAB_TIPO_DOCUMENTO_NCDOC.value[0])
        ws.cell(row=next_row, column=Cabecera.ENCAB_NO_DTO_EXT.value[1], value=item["FacturaRelacionada"])
    
    # Rellenar las demás celdas
    ws.cell(row=next_row, column=Constants.ENCAB_TERCERO_INTERNO.value[1], value=Constants.ENCAB_TERCERO_INTERNO.value[0])
    ws.cell(row=next_row, column=Constants.ENCAB_TERCERO_EXTERNO.value[1], value=Constants.ENCAB_TERCERO_EXTERNO.value[0])
    ws.cell(row=next_row, column=Constants.ENCAB_FORMA_PAGO.value[1], value=Constants.ENCAB_FORMA_PAGO.value[0])
    ws.cell(row=next_row, column=Constants.ENCAB_VERIFICADO.value[1], value=Constants.ENCAB_VERIFICADO.value[0])
    ws.cell(row=next_row, column=Constants.ENCAB_ANULADO.value[1], value=Constants.ENCAB_ANULADO.value[0])
    ws.cell(row=next_row, column=Constants.DETALLE_PRODUCTO.value[1], value=Constants.DETALLE_PRODUCTO.value[0])
    ws.cell(row=next_row, column=Constants.DETALLE_BODEGA.value[1], value=Constants.DETALLE_BODEGA.value[0])
    ws.cell(row=next_row, column=Constants.DETALLE_UNIDAD_MEDIDA.value[1], value=Constants.DETALLE_UNIDAD_MEDIDA.value[0])
    ws.cell(row=next_row, column=Constants.DETALLE_CANTIDAD.value[1], value=Constants.DETALLE_CANTIDAD.value[0])
    ws.cell(row=next_row, column=Constants.DETALLE_IVA.value[1], value=Constants.DETALLE_IVA.value[0])
    ws.cell(row=next_row, column=Constants.DETALLE_DESCUENTO.value[1], value=Constants.DETALLE_DESCUENTO.value[0])

    # Rellenar los campos específicos del item
    ws.cell(row=next_row, column=Cabecera.ENCAB_FECHA.value[1], value=item["FechaEmision"])
    ws.cell(row=next_row, column=Cabecera.ENCAB_PREF_DTO_EXT.value[1], value=item["FacturaCabecera"])
    ws.cell(row=next_row, column=Cabecera.ENCAB_NOTA.value[1], value=item["NombrePeaje"])
    ws.cell(row=next_row, column=Cabecera.ENCAB_FECHA_EMISION.value[1], value=item["FechaEmision"])

    ws.cell(row=next_row, column=Cabecera.DETALLE_VALOR_UNITARIO.value[1], value=item["ValorTotal"])
    ws.cell(row=next_row, column=Cabecera.DETALLE_VENCIMIENTO.value[1], value=item["FechaEmision"])
    ws.cell(row=next_row, column=Cabecera.DETALLE_CENTRO_COSTOS.value[1], value=item["NumeroPlaca"])



